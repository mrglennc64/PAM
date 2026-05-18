#!/usr/bin/env python3
import argparse
import csv
import datetime as dt
import hashlib
import json
import os
import sys
from pathlib import Path

SKILL_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SKILL_DIR))

from checks import isrc_iswc, role_splits, duplicates, reference_match

OUTBOX = Path("/opt/ai-agency/music-catalog/outbox")
ARCHIVE = Path("/opt/ai-agency/music-catalog/archive")
HUMAN_REVIEW_THRESHOLD = 0.85

FORBIDDEN_TERMS = ("royalty", "royalties", "mechanical rate", "per stream", "earnings")


def _refuse_if_financial(prompt):
    if not prompt:
        return
    low = prompt.lower()
    for term in FORBIDDEN_TERMS:
        if term in low:
            print(json.dumps({
                "status": "failed",
                "reason": "no_financial_calculations",
                "detail": f"Skill refuses tasks involving financial computation (matched: {term!r})",
            }))
            sys.exit(2)


def _read_rows(path):
    p = Path(path)
    if not p.is_file():
        sys.exit(f"input not found: {path}")
    if p.suffix.lower() == ".csv":
        with open(p, "r", newline="", encoding="utf-8") as f:
            return list(csv.DictReader(f))
    if p.suffix.lower() in (".xlsx", ".xls"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            sys.exit("openpyxl required for xlsx input — `pip install openpyxl`")
        wb = load_workbook(p, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = [str(c) if c is not None else "" for c in next(rows_iter)]
        return [dict(zip(header, [("" if c is None else c) for c in row])) for row in rows_iter]
    sys.exit(f"unsupported format: {p.suffix}")


def _md5(path):
    h = hashlib.md5()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def main():
    parser = argparse.ArgumentParser(description="music-catalog-scan")
    parser.add_argument("--input", required=True, help="path to CSV/XLSX in inbox/")
    parser.add_argument("--prompt", default="", help="originating task prompt (for financial-term guard)")
    parser.add_argument("--archive", action="store_true", help="move input to archive/ after scan")
    args = parser.parse_args()

    _refuse_if_financial(args.prompt)

    input_path = Path(args.input).resolve()
    pre_md5 = _md5(input_path)

    rows = _read_rows(input_path)
    findings = []

    for row in rows:
        findings.extend(isrc_iswc.check(row))
        findings.extend(role_splits.check(row))
        findings.extend(reference_match.check(row))

    findings.extend(duplicates.check_batch(rows))

    OUTBOX.mkdir(parents=True, exist_ok=True)
    stamp = dt.datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
    out_path = OUTBOX / f"{input_path.stem}-discrepancies-{stamp}.jsonl"
    with open(out_path, "w", encoding="utf-8") as f:
        for r in findings:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")

    post_md5 = _md5(input_path)
    if pre_md5 != post_md5:
        sys.exit(f"FATAL: input file modified during scan (pre={pre_md5} post={post_md5}). Aborting.")

    requires_review = any(
        f.get("severity") == "error" or f.get("confidence", 1.0) < HUMAN_REVIEW_THRESHOLD
        for f in findings
    )

    summary = {
        "status": "review" if requires_review else "done",
        "input": str(input_path),
        "input_md5": pre_md5,
        "output": str(out_path),
        "rows_scanned": len(rows),
        "findings": len(findings),
        "errors": sum(1 for f in findings if f.get("severity") == "error"),
        "warnings": sum(1 for f in findings if f.get("severity") == "warning"),
        "requires_human_review": requires_review,
    }

    if args.archive:
        ARCHIVE.mkdir(parents=True, exist_ok=True)
        dest = ARCHIVE / f"{input_path.stem}-{stamp}{input_path.suffix}"
        os.replace(input_path, dest)
        summary["archived_to"] = str(dest)

    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
