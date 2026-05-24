"""
Google Sheets Template Generator
Converts any probe-ready CSV into a formatted .xlsx with:
  - Clickable HYPERLINK formulas (ISRC, Name Search, YouTube)
  - Conditional formatting rules (JACKPOT=green, etc.)
  - Frozen header row + auto column widths

Usage:
    pip install openpyxl
    python probe/make_sheets_template.py data/female_rappers_2024_present.csv
    python probe/make_sheets_template.py data/gospel_soul_targets.csv
    python probe/make_sheets_template.py data/atl_remix_targets_2025.csv
"""

import sys, os, csv
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

input_path = sys.argv[1] if len(sys.argv) > 1 else "data/female_rappers_2024_present.csv"
base = os.path.splitext(os.path.basename(input_path))[0]
output_path = os.path.join(os.path.dirname(input_path), f"{base}_template.xlsx")

# Load CSV
with open(input_path, newline="", encoding="utf-8") as f:
    rows = list(csv.DictReader(f))

# Normalize columns
def norm(d):
    return {k.strip().lower().replace(" ","_"): (v or "").strip() for k,v in d.items()}
rows = [norm(r) for r in rows]
if rows and "song" in rows[0] and "track" not in rows[0]:
    for r in rows:
        r["track"] = r.pop("song", "")

# ── Workbook setup ──────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Probe List"

# ── Colors ──────────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="0F172A")   # dark navy
HEADER_FONT   = Font(name="Calibri", bold=True, color="F8FAFC", size=11)
GREEN_FILL    = PatternFill("solid", fgColor="00C851")
LT_GREEN_FILL = PatternFill("solid", fgColor="CCFFDD")
RED_FILL      = PatternFill("solid", fgColor="EF4444")
YELLOW_FILL   = PatternFill("solid", fgColor="FEF08A")
ORANGE_FILL   = PatternFill("solid", fgColor="FB923C")
BORDER_SIDE   = Side(style="thin", color="CBD5E1")
CELL_BORDER   = Border(bottom=BORDER_SIDE)

# ── Headers ─────────────────────────────────────────────────────────────────
headers = [
    "Artist", "Track", "ISRC", "Sniper Score",
    "SX ISRC Link", "SX Name Search", "MusicBrainz", "Songview (ASCAP+BMI)",
    "YouTube", "Release Date", "BB Eligibility", "Manual Check Notes", "Estimated Recovery"
]
ws.append(headers)
for col, _ in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col)
    cell.fill   = HEADER_FILL
    cell.font   = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ── Data rows ───────────────────────────────────────────────────────────────
for i, r in enumerate(rows, start=2):
    artist  = r.get("artist", "")
    track   = r.get("track",  "")
    isrc    = r.get("isrc",   "")
    score   = r.get("sniper_score", 0)

    # Col A-D: raw values
    ws.cell(row=i, column=1, value=artist)
    ws.cell(row=i, column=2, value=track)
    ws.cell(row=i, column=3, value=isrc if isrc.lower() not in ("nan","","none") else "")
    ws.cell(row=i, column=4, value=int(score) if str(score).isdigit() else 0)

    # Col E: SX ISRC link — only if ISRC present
    if isrc and isrc.lower() not in ("nan","","none"):
        ws.cell(row=i, column=5,
            value=f'=HYPERLINK("https://isrc.soundexchange.com/#/search?searchType=code&query={isrc}","🔍 ISRC Search")')
    else:
        ws.cell(row=i, column=5, value="—")

    # Col F: SX Name Search
    ws.cell(row=i, column=6,
        value=f'=HYPERLINK("https://isrc.soundexchange.com/#/search?searchType=name&query="&ENCODEURL(A{i}&" "&B{i}),"🔍 Name Search")')

    # Col G: MusicBrainz
    if isrc and isrc.lower() not in ("nan","","none"):
        ws.cell(row=i, column=7,
            value=f'=HYPERLINK("https://musicbrainz.org/isrc/{isrc}","🎵 MusicBrainz")')
    else:
        ws.cell(row=i, column=7,
            value=f'=HYPERLINK("https://musicbrainz.org/search?query="&ENCODEURL(A{i}&" "&B{i})&"&type=recording","🎵 MusicBrainz")')

    # Col H: Songview (ASCAP+BMI)
    ws.cell(row=i, column=8,
        value=f'=HYPERLINK("https://repertoire.bmi.com/search?title="&ENCODEURL(B{i})&"&searchIn=songview","📋 Songview")')

    # Col I: YouTube
    ws.cell(row=i, column=9,
        value=f'=HYPERLINK("https://www.youtube.com/results?search_query="&ENCODEURL(A{i}&" "&B{i}),"▶ YouTube")')

    # Col J: Release Date
    ws.cell(row=i, column=10, value=r.get("release_date", ""))

    # Col K: Black Box Eligibility
    elig = r.get("eligibility", "")
    elig_cell = ws.cell(row=i, column=11, value=elig)
    if "FULL" in elig:
        elig_cell.fill = PatternFill("solid", fgColor="DCFCE7")
        elig_cell.font = Font(bold=True, color="166534")
    elif "PARTIAL" in elig:
        elig_cell.fill = PatternFill("solid", fgColor="FEF9C3")
        elig_cell.font = Font(bold=True, color="854D0E")
    elif "CHECK" in elig:
        elig_cell.fill = PatternFill("solid", fgColor="FEE2E2")
        elig_cell.font = Font(color="991B1B")

    # Col L: Manual Check Notes (empty — user fills in)
    ws.cell(row=i, column=12, value="")

    # Col M: Estimated Recovery (empty)
    ws.cell(row=i, column=13, value="")

    # Row styling
    for col in range(1, 12):
        cell = ws.cell(row=i, column=col)
        cell.border    = CELL_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=False)
        if i % 2 == 0:
            cell.fill = PatternFill("solid", fgColor="F1F5F9")

# ── Conditional Formatting (column J = Manual Check Notes) ──────────────────
last_row = len(rows) + 1
notes_range  = f"L2:L{last_row}"
score_range  = f"D2:D{last_row}"

# JACKPOT → green
ws.conditional_formatting.add(notes_range,
    FormulaRule(formula=[f'NOT(ISERROR(SEARCH("JACKPOT",J2)))'],
                fill=GREEN_FILL, font=Font(bold=True, color="FFFFFF")))

# GUEST_UNCLAIMED → light green
ws.conditional_formatting.add(notes_range,
    FormulaRule(formula=[f'NOT(ISERROR(SEARCH("GUEST_UNCLAIMED",J2)))'],
                fill=LT_GREEN_FILL, font=Font(bold=True)))

# False Positive → red
ws.conditional_formatting.add(notes_range,
    FormulaRule(formula=[f'NOT(ISERROR(SEARCH("False Positive",J2)))'],
                fill=RED_FILL, font=Font(color="FFFFFF")))

# Needs Check → yellow
ws.conditional_formatting.add(notes_range,
    FormulaRule(formula=[f'NOT(ISERROR(SEARCH("Needs Check",J2)))'],
                fill=YELLOW_FILL))

# Sniper Score > 80 → orange
ws.conditional_formatting.add(score_range,
    CellIsRule(operator="greaterThan", formula=["80"],
               fill=ORANGE_FILL, font=Font(bold=True)))

# ── Column widths ────────────────────────────────────────────────────────────
col_widths = [22, 38, 16, 13, 14, 15, 14, 20, 12, 13, 22, 22, 18]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── Freeze header row ────────────────────────────────────────────────────────
ws.freeze_panes = "A2"

# ── Row height ───────────────────────────────────────────────────────────────
ws.row_dimensions[1].height = 32
for i in range(2, last_row + 1):
    ws.row_dimensions[i].height = 20

wb.save(output_path)
print(f"Template created: {output_path}")
print(f"  {len(rows)} rows | Upload to Google Sheets or open in Excel")
print(f"  Conditional formatting: JACKPOT=green, GUEST_UNCLAIMED=light green,")
print(f"  False Positive=red, Needs Check=yellow, Score>80=orange")
